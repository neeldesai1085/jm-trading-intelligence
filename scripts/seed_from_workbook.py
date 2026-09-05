import sys
from pathlib import Path
sys.path.insert(0,str(Path(__file__).resolve().parents[1]/'backend'))
from openpyxl import load_workbook
from sqlalchemy import select
from app.db.session import Base,engine,SessionLocal
from app.models.entities import User,ContractNote,SecurityLedger,Execution,Portfolio
from app.services.auth import hash_password

def main(path,email='demo@example.com',password='DemoPass123!'):
    Base.metadata.create_all(bind=engine);wb=load_workbook(path,read_only=True,data_only=True);db=SessionLocal()
    try:
        user=db.scalar(select(User).where(User.email==email.lower()))
        if not user:user=User(email=email.lower(),name='Demo Trader',password_hash=hash_password(password),is_active=True);db.add(user);db.commit();db.refresh(user)
        portfolio=db.scalar(select(Portfolio).where(Portfolio.user_id==user.id,Portfolio.is_default.is_(True)))
        if not portfolio:portfolio=Portfolio(user_id=user.id,name='Main Portfolio',is_default=True);db.add(portfolio);db.commit();db.refresh(portfolio)
        for model in [ContractNote,SecurityLedger,Execution]:db.query(model).filter(model.user_id==user.id,model.portfolio_id==portfolio.id).delete(synchronize_session=False)
        ws=wb['Contract Notes'];rows=list(ws.iter_rows(values_only=True));hdr=rows[0]
        for row in rows[1:]:
            d=dict(zip(hdr,row));dt=d['Trade Date'].date() if hasattr(d['Trade Date'],'date') else d['Trade Date'];sd=d['Settlement Date'].date() if hasattr(d['Settlement Date'],'date') else d['Settlement Date'];db.add(ContractNote(user_id=user.id,portfolio_id=portfolio.id,contract_note=str(d['Contract Note']),trade_date=dt,settlement_date=sd,buy_qty=int(d['Buy Qty'] or 0),sell_qty=int(d['Sell Qty'] or 0),gross_buy_value=float(d['Gross Buy Value'] or 0),gross_sell_value=float(d['Gross Sell Value'] or 0),displayed_brokerage=float(d['Displayed Brokerage'] or 0),buy_value_after_brokerage=float(d['Buy Value After Brokerage'] or 0),sell_value_after_brokerage=float(d['Sell Value After Brokerage'] or 0),market_flow_after_brokerage=float(d['Market Flow After Brokerage'] or 0),payin_obligation=float(d['Pay-in/Pay-out Obligation'] or 0),taxable_value=float(d['Taxable Value of Supply'] or 0),stt=float(d['STT'] or 0),cgst=float(d['CGST'] or 0),sgst=float(d['SGST'] or 0),ugst=float(d['UGST'] or 0),igst=float(d['IGST'] or 0),exchange_charges=float(d['Exchange Transaction Charges'] or 0),sebi_fees=float(d['SEBI Turnover Fees'] or 0),stamp_duty=float(d['Stamp Duty'] or 0),ipft=float(d['IPFT Charges'] or 0),net_amount=float(d['Net Amount'] or 0),contract_note_page=None,annexure_page=None,settlement_no=None,source_file=str(path)))
        ws=wb['Security Ledger'];rows=list(ws.iter_rows(values_only=True));hdr=rows[0]
        for row in rows[1:]:
            d=dict(zip(hdr,row));td=d['Trade Date'].date() if hasattr(d['Trade Date'],'date') else d['Trade Date'];db.add(SecurityLedger(user_id=user.id,portfolio_id=portfolio.id,contract_note=str(d['Contract Note']),trade_date=td,isin=d['ISIN'],security=d['Security'],buy_qty=int(d['Buy Qty'] or 0),buy_wap=float(d['Buy WAP'] or 0),buy_brokerage_share=float(d['Buy Brokerage/Share'] or 0),buy_wap_after_brokerage=float(d['Buy WAP After Brokerage'] or 0),total_buy_value_after_brokerage=float(d['Total Buy Value After Brokerage'] or 0),gross_buy=float(d['Gross Buy'] or 0),displayed_buy_brokerage=float(d['Displayed Buy Brokerage'] or 0),sell_qty=int(d['Sell Qty'] or 0),sell_wap=float(d['Sell WAP'] or 0),sell_brokerage_share=float(d['Sell Brokerage/Share'] or 0),sell_wap_after_brokerage=float(d['Sell WAP After Brokerage'] or 0),total_sell_value_after_brokerage=float(d['Total Sell Value After Brokerage'] or 0),gross_sell=float(d['Gross Sell'] or 0),displayed_sell_brokerage=float(d['Displayed Sell Brokerage'] or 0),net_qty=int(d['Net Qty'] or 0),net_obligation_before_levies=float(d['Net Obligation Before Levies'] or 0)))
        ws=wb['Execution Ledger'];rows=list(ws.iter_rows(values_only=True));hdr=rows[0]
        for row in rows[1:]:
            d=dict(zip(hdr,row));td=d['Trade Date'].date() if hasattr(d['Trade Date'],'date') else d['Trade Date'];db.add(Execution(user_id=user.id,portfolio_id=portfolio.id,contract_note=str(d['Contract Note']),trade_date=td,order_no=str(d['Order No']),order_time=str(d['Order Time']),trade_no=str(d['Trade No']),trade_time=str(d['Trade Time']),security=d['Security'],exchange=d['Exchange'],side=d['Side'],quantity=int(d['Quantity']),market_rate=float(d['Market Rate']),amount=float(d['Amount'])))
        db.commit()
    finally:db.close();wb.close()
if __name__=='__main__':main(sys.argv[1],*(sys.argv[2:]))
